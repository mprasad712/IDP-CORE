"""Turn a built document payload (or generic tabular rows) into downloadable bytes.

Pure functions — no DB, no FastAPI. One serializer per format for a single document
(JSON/CSV/XLSX/XML/TXT), plus generic ``tabular_*`` helpers (column list + list-of-dict
rows) reused by the date-range report exports.

``None`` becomes an empty cell in CSV/XLSX/XML; only JSON keeps ``None``/``""`` distinct
(ElementTree renders both ``text=""`` and ``text=None`` as ``<tag/>``, and CSV/XLSX have
no null — so the audited-empty-vs-never-reviewed distinction lives in JSON + the payload).
"""
from __future__ import annotations

import csv
import io
import json
import xml.etree.ElementTree as ET
from typing import Any, Iterable, Sequence

import openpyxl

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _scalar(v: Any):
    """Normalize a value for a flat cell: None→"", bool→'true'/'false', else unchanged."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"
    return v


def _xml_tag(name: Any) -> str:
    """Make a safe XML tag from a column name (letters/digits/underscore; never empty)."""
    safe = "".join(ch if (ch.isalnum() or ch == "_") else "_" for ch in str(name))
    if not safe or not (safe[0].isalpha() or safe[0] == "_"):
        safe = f"_{safe}"
    return safe


def _disp(v: Any) -> str:
    """Human-readable rendering for the TXT format: None→em-dash, ""→(blank)."""
    if v is None:
        return "—"
    if v == "":
        return "(blank)"
    return str(v)


# ───────────────────────────── single-document serializers ─────────────────────────────

def to_json(payload: dict) -> bytes:
    """Nested JSON — preserves None vs "" and the full predicted/audited structure."""
    return json.dumps(payload, indent=2, ensure_ascii=False, default=str).encode("utf-8")


def to_csv(payload: dict) -> bytes:
    """Document-meta key/value block, then a Header Fields table, then a Line Items table."""
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Document"])
    for k, v in payload["document"].items():
        w.writerow([k, _scalar(v)])
    w.writerow([])
    w.writerow(["Header Fields"])
    hcols = payload["header_columns"]
    w.writerow(hcols)
    for row in payload["headers"]:
        w.writerow([_scalar(row.get(c)) for c in hcols])
    w.writerow([])
    w.writerow(["Line Items"])
    lcols = payload["line_item_columns"]
    w.writerow(lcols)
    for row in payload["line_items"]:
        w.writerow([_scalar(row.get(c)) for c in lcols])
    # utf-8-sig → Excel opens the CSV with the correct encoding (BOM); new file, no regression.
    return out.getvalue().encode("utf-8-sig")


def to_xlsx(payload: dict) -> bytes:
    """Three sheets: Summary (document meta) · Headers · Line Items.

    A single document is small, so this builds the workbook in memory (plain ``Workbook()``,
    no ``write_only`` temp file) — more robust on hosts with a restricted/unwritable TMPDIR.
    The BULK report export uses ``tabular_to_xlsx`` (write_only) where the row count actually
    warrants streaming to disk.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    for k, v in payload["document"].items():
        ws.append([k, _scalar(v)])
    hs = wb.create_sheet("Headers")
    hcols = payload["header_columns"]
    hs.append(list(hcols))
    for row in payload["headers"]:
        hs.append([_scalar(row.get(c)) for c in hcols])
    ls = wb.create_sheet("Line Items")
    lcols = payload["line_item_columns"]
    ls.append(list(lcols))
    for row in payload["line_items"]:
        ls.append([_scalar(row.get(c)) for c in lcols])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def to_xml(payload: dict) -> bytes:
    """``<document><meta/><headers><field/>…</headers><line_items><row/>…</line_items></document>``."""
    root = ET.Element("document")
    meta = ET.SubElement(root, "meta")
    for k, v in payload["document"].items():
        el = ET.SubElement(meta, _xml_tag(k))
        el.text = None if v is None else str(v)
    headers_el = ET.SubElement(root, "headers")
    for row in payload["headers"]:
        fel = ET.SubElement(headers_el, "field")
        for c in payload["header_columns"]:
            cel = ET.SubElement(fel, _xml_tag(c))
            val = row.get(c)
            cel.text = None if val is None else str(val)
    lines_el = ET.SubElement(root, "line_items")
    for row in payload["line_items"]:
        rel = ET.SubElement(lines_el, "row")
        for c in payload["line_item_columns"]:
            cel = ET.SubElement(rel, _xml_tag(c))
            val = row.get(c)
            cel.text = None if val is None else str(val)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _field_line(row: dict, final: bool) -> str:
    conf = row.get("confidence")
    conf_s = f"{conf}%" if conf is not None else "n/a"
    if final:
        return f"{_disp(row.get('value'))} (conf {conf_s})"
    aud = _disp(row.get("audited_value")) if row.get("is_reviewed") else "<not reviewed>"
    return f"predicted={_disp(row.get('predicted_value'))} | audited={aud} (conf {conf_s})"


def to_txt(payload: dict) -> bytes:
    """Human-readable plain-text rendering."""
    doc = payload["document"]
    final = payload.get("values_mode") == "final"
    out: list[str] = []
    out.append(f"Document : {doc.get('filename') or '—'}")
    out.append(f"ID       : {doc.get('id') or '—'}")
    out.append(f"Type     : {doc.get('predicted_type') or '—'}")
    out.append(f"Status   : {doc.get('status') or '—'}")
    oc = doc.get("overall_confidence")
    out.append(f"Confidence: {oc}%" if oc is not None else "Confidence: —")
    out.append(f"Uploaded : {doc.get('uploaded_at') or '—'}")
    out.append(f"Processed: {doc.get('pipeline_completed_at') or '—'}")
    out.append(f"Reviewed : {doc.get('reviewed_at') or '—'} ({doc.get('reviewer') or 'no reviewer'})")
    out.append("")
    out.append("HEADER FIELDS")
    out.append("-" * 48)
    for row in payload["headers"]:
        out.append(f"  {row.get('field_name')}: {_field_line(row, final)}")
    out.append("")
    out.append("LINE ITEMS")
    out.append("-" * 48)
    last = object()
    for row in payload["line_items"]:
        ri = row.get("row_index")
        if ri != last:
            out.append(f"  Row {ri}:")
            last = ri
        out.append(f"    {row.get('column_name')}: {_field_line(row, final)}")
    return ("\n".join(out) + "\n").encode("utf-8")


# ───────────────────────── generic tabular serializers (reports) ─────────────────────────

def tabular_to_csv(columns: Sequence[str], rows: Iterable[dict]) -> bytes:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(list(columns))
    for r in rows:
        w.writerow([_scalar(r.get(c)) for c in columns])
    return out.getvalue().encode("utf-8-sig")


def tabular_to_xlsx(columns: Sequence[str], rows: Iterable[dict], *, sheet_name: str = "Data") -> bytes:
    # write_only keeps memory bounded for large exports (rows streamed straight to the file).
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title=str(sheet_name)[:31])
    ws.append(list(columns))
    for r in rows:
        ws.append([_scalar(r.get(c)) for c in columns])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def tabular_to_xml(
    columns: Sequence[str], rows: Iterable[dict], *, root_tag: str = "report", row_tag: str = "record"
) -> bytes:
    root = ET.Element(root_tag)
    for r in rows:
        rec = ET.SubElement(root, row_tag)
        for c in columns:
            cel = ET.SubElement(rec, _xml_tag(c))
            val = r.get(c)
            cel.text = None if val is None else str(val)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


# ───────────────── matrix serializers (positional grids — sectioned exports) ─────────────────
# A "matrix" is a list of rows, each row a list of raw cell values (ragged rows allowed). Used
# for the sectioned "Download all data" layout where each file is its own block (file-info +
# header table beside a line-items table), so there is no single uniform column set.

def matrix_to_csv(matrix: Iterable[Sequence]) -> bytes:
    out = io.StringIO()
    w = csv.writer(out)
    for row in matrix:
        w.writerow([_scalar(c) for c in row])
    return out.getvalue().encode("utf-8-sig")


def matrix_to_xlsx(matrix: Iterable[Sequence], *, sheet_name: str = "Data") -> bytes:
    # write_only streams rows to disk → bounded memory for large stacked exports.
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title=str(sheet_name)[:31])
    for row in matrix:
        ws.append([_scalar(c) for c in row])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def matrix_to_xml(
    matrix: Iterable[Sequence], *, root_tag: str = "data", row_tag: str = "row", cell_tag: str = "c"
) -> bytes:
    root = ET.Element(root_tag)
    for row in matrix:
        rel = ET.SubElement(root, row_tag)
        for c in row:
            cel = ET.SubElement(rel, cell_tag)
            cel.text = None if c is None or c == "" else str(c)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)
