"""Unit tests for the IDP output serialization module (pure functions, no DB).

Covers audit-column honesty (predicted vs audited vs is_reviewed; "" vs None), the
values=final collapse, and that every format produces parseable bytes with the right
media type / extension.
"""
import io
import json
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from types import SimpleNamespace
from uuid import uuid4

import openpyxl
import pytest

from agentcore.services.idp.output import (
    SUPPORTED_FORMATS,
    serialize_document,
    serialize_table,
)
from agentcore.services.idp.output.schema import build_doc_payload


def _doc():
    return SimpleNamespace(
        id=uuid4(),
        original_filename="invoice_42.pdf",
        predicted_type="invoice",
        status="reviewed",
        overall_confidence=0.875,
        created_at=datetime(2026, 6, 1, 10, 0, tzinfo=timezone.utc),
        processing_completed_at=datetime(2026, 6, 1, 10, 1, tzinfo=timezone.utc),
    )


def _header(field_name, extracted, reviewed, is_reviewed, conf):
    return SimpleNamespace(
        field_name=field_name, extracted_value=extracted, reviewed_value=reviewed,
        is_reviewed=is_reviewed, confidence_score=conf,
    )


def _line(row_index, column_name, extracted, reviewed, is_reviewed, conf):
    return SimpleNamespace(
        row_index=row_index, column_name=column_name, extracted_value=extracted,
        reviewed_value=reviewed, is_reviewed=is_reviewed, confidence_score=conf,
    )


def _sample(values="both"):
    doc = _doc()
    headers = [
        _header("vendor", "Acme", "Acme Corp", True, 0.9),   # corrected by a human
        _header("total", "100.00", None, False, 0.5),         # never reviewed → audited None
        _header("po_num", "PO-1", "", True, None),            # human blanked it → audited "" (kept!)
    ]
    lines = [
        _line(1, "item", "Widget", None, False, 0.8),
        _line(1, "qty", "2", "3", True, 0.7),
    ]
    rs = SimpleNamespace(
        review_completed_at=datetime(2026, 6, 2, 9, 0, tzinfo=timezone.utc), reviewed_by=uuid4()
    )
    return build_doc_payload(doc, headers, lines, rs, values=values)


def test_payload_audit_honesty_both():
    p = _sample("both")
    h = {r["field_name"]: r for r in p["headers"]}
    assert h["vendor"]["predicted_value"] == "Acme"
    assert h["vendor"]["audited_value"] == "Acme Corp"
    assert h["vendor"]["is_reviewed"] is True
    # never-reviewed → audited None (NOT "")
    assert h["total"]["audited_value"] is None
    assert h["total"]["is_reviewed"] is False
    # human-blanked → audited "" preserved distinct from None
    assert h["po_num"]["audited_value"] == ""
    assert h["po_num"]["is_reviewed"] is True
    # confidence rendered as percent
    assert h["vendor"]["confidence"] == 90.0
    assert h["total"]["confidence"] == 50.0
    assert h["po_num"]["confidence"] is None
    # document meta
    assert p["document"]["overall_confidence"] == 87.5
    assert p["document"]["pipeline_completed_at"] == "2026-06-01T10:01:00+00:00"
    assert p["document"]["reviewed_at"] == "2026-06-02T09:00:00+00:00"
    assert p["document"]["reviewer"] is not None


def test_payload_values_final_collapse():
    p = _sample("final")
    h = {r["field_name"]: r for r in p["headers"]}
    assert p["header_columns"] == ["field_name", "value", "confidence"]
    assert h["vendor"]["value"] == "Acme Corp"   # reviewed → audited value
    assert h["total"]["value"] == "100.00"        # not reviewed → predicted value
    assert h["po_num"]["value"] == ""             # reviewed-blank → ""
    assert "predicted_value" not in h["vendor"]


def test_json_preserves_none_vs_blank():
    data, media, ext = serialize_document(_sample("both"), "json")
    assert media == "application/json" and ext == "json"
    obj = json.loads(data)
    h = {r["field_name"]: r for r in obj["headers"]}
    assert h["total"]["audited_value"] is None
    assert h["po_num"]["audited_value"] == ""


def test_csv_has_sections_and_rows():
    data, media, ext = serialize_document(_sample("both"), "csv")
    assert media == "text/csv" and ext == "csv"
    text = data.decode("utf-8-sig")
    assert "Header Fields" in text and "Line Items" in text
    assert "vendor" in text and "Acme Corp" in text


def test_xlsx_three_sheets():
    data, media, ext = serialize_document(_sample("both"), "excel")
    assert ext == "xlsx"
    assert media.endswith("spreadsheetml.sheet")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Summary", "Headers", "Line Items"]


def test_xml_parses():
    data, media, ext = serialize_document(_sample("both"), "xml")
    assert media == "application/xml" and ext == "xml"
    root = ET.fromstring(data)
    assert root.tag == "document"
    fields = root.find("headers").findall("field")
    assert len(fields) == 3


def test_txt_human_readable():
    data, media, ext = serialize_document(_sample("both"), "txt")
    assert ext == "txt"
    text = data.decode("utf-8")
    assert "HEADER FIELDS" in text and "LINE ITEMS" in text
    assert "invoice_42.pdf" in text
    assert "<not reviewed>" in text   # the 'total' field was never reviewed


def test_unknown_format_raises():
    with pytest.raises(ValueError):
        serialize_document(_sample("both"), "pdf")


def test_supported_formats_constant():
    assert SUPPORTED_FORMATS == {"csv", "excel", "xml", "json", "txt"}


def test_build_doc_payload_rejects_bad_values():
    with pytest.raises(ValueError):
        build_doc_payload(_doc(), [], [], None, values="garbage")


def test_serialize_table_roundtrip():
    cols = ["a", "b", "flag"]
    rows = [{"a": 1, "b": None, "flag": True}, {"a": 2, "b": "x", "flag": False}]
    csv_bytes, media, ext = serialize_table(cols, rows, "csv")
    assert ext == "csv"
    first = csv_bytes.decode("utf-8-sig").splitlines()[0]
    assert first == "a,b,flag"
    xlsx_bytes, _m, xext = serialize_table(cols, rows, "excel", sheet_name="Report")
    assert xext == "xlsx"
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.sheetnames == ["Report"]
    xml_bytes = serialize_table(cols, rows, "xml", root_tag="report", row_tag="po")[0]
    root = ET.fromstring(xml_bytes)
    assert root.tag == "report" and len(root.findall("po")) == 2


def test_serialize_table_rejects_bad_format():
    with pytest.raises(ValueError):
        serialize_table(["a"], [{"a": 1}], "json")   # json/txt not offered for bulk tables
